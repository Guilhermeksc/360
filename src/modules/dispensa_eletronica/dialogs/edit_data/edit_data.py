from PyQt6.QtWidgets import *
from PyQt6.QtGui import *
from PyQt6.QtCore import *
from src.modules.utils.add_button import add_button, add_button_func
from src.modules.dispensa_eletronica.dialogs.edit_data.widgets.consulta_api import setup_consulta_api
from src.modules.dispensa_eletronica.dialogs.edit_data.widgets.gerar_documentos import ConsolidarDocumentos
from src.modules.dispensa_eletronica.dialogs.edit_data.widgets.sigdem_layout import create_GrupoSIGDEM, create_utilidades_group
from src.modules.dispensa_eletronica.dialogs.edit_data.widgets.setor_responsavel import create_dados_responsavel_contratacao_group
from src.modules.dispensa_eletronica.dialogs.edit_data.widgets.contratacao import create_info_comprasnet
from src.modules.utils.linha_layout import linha_divisoria_layout
from src.modules.utils.select_om import create_selecao_om_layout, load_sigla_om, on_om_changed
from src.modules.utils.agentes_responsaveis_layout import create_combo_box, carregar_agentes_responsaveis
from pathlib import Path
from src.config.paths import CONTROLE_DADOS
import json
import pandas as pd
import os
import subprocess
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from src.modules.utils.add_button import add_button, add_button_func

CONFIG_FILE = 'config.json'

def load_config_path_id():
    if not Path(CONFIG_FILE).exists():
        return {}
    with open(CONFIG_FILE, 'r') as file:
        return json.load(file)

def save_config(config):
    with open(CONFIG_FILE, 'w') as file:
        json.dump(config, file)
from PyQt6.QtCore import QThread, pyqtSignal
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from pathlib import Path
import os
import subprocess

def number_to_text(number):
    numbers_in_words = ["um", "dois", "três", "quatro", "cinco", "seis", "sete", "oito", "nove", "dez", "onze", "doze"]
    return numbers_in_words[number - 1] 

class TableCreationWorker(QThread):
    # Sinal para indicar quando o arquivo foi salvo
    file_saved = pyqtSignal(str)

    def __init__(self, dados, colunas_legiveis, pasta_base):
        super().__init__()
        self.dados = dados
        self.colunas_legiveis = colunas_legiveis
        self.pasta_base = pasta_base

    def run(self):
        # Criando e salvando a tabela
        wb = Workbook()
        ws = wb.active
        ws.title = "Formulário"
        
        # Adicionando título
        tipo = self.dados.get('tipo', '')
        numero = self.dados.get('numero', '')
        ano = self.dados.get('ano', '')
        objeto = self.dados.get('objeto', '')
        titulo = f"{tipo} nº {numero}/{ano} ({objeto})"
        ws.merge_cells('A1:B1')
        ws['A1'] = titulo
        ws['A1'].font = Font(size=20, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 40

        # Cabeçalhos
        ws['A2'] = "Índice"
        ws['B2'] = "Valor"
        ws['A2'].font = Font(size=14, bold=True)
        ws['B2'].font = Font(size=14, bold=True)
        ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
        ws['B2'].alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        ws['A2'].border = thin_border
        ws['B2'].border = thin_border

        # Ajuste de largura das colunas
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 80

        # Preenchendo dados filtrados
        for i, (key, label) in enumerate(self.colunas_legiveis.items(), start=3):
            value = self.dados.get(key, '')  # Obtém o valor correspondente em `self.dados`
            ws[f'A{i}'] = label
            ws[f'B{i}'] = str(value)
            ws[f'B{i}'].alignment = Alignment(wrap_text=True)
            fill_color = "F2F2F2" if i % 2 == 0 else "FFFFFF"
            fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            ws[f'A{i}'].fill = fill
            ws[f'B{i}'].fill = fill
            ws[f'A{i}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.row_dimensions[i].height = 15

            # Aplicando bordas finas a cada célula preenchida
            for cell in [ws[f'A{i}'], ws[f'B{i}']]:
                cell.border = thin_border

        # Salvando arquivo e emitindo o sinal com o caminho
        file_path = self.pasta_base / "formulario.xlsx"
        wb.save(file_path)
        self.file_saved.emit(str(file_path))  # Emite o sinal com o caminho do arquivo salvo


class EditarDadosWindow(QMainWindow):
    save_data_signal = pyqtSignal(dict)
    status_atualizado = pyqtSignal(str, str)
    formulario_carregado = pyqtSignal(pd.DataFrame) 

    def __init__(self, dados, icons, parent=None):
        super().__init__(parent)
        self.dados = dados
        self.icons = icons

        # Configurações gerais da janela
        self.setWindowTitle("Editar Dados")
        self.setWindowIcon(self.icons.get("edit", None))
        self.setFixedSize(1150, 780)

        # Referências aos RadioButtons para material_servico, com_disputa e pesquisa_preco
        self.radio_material = None
        self.radio_servico = None
        self.radio_disputa_sim = None
        self.radio_disputa_nao = None
        self.radio_pesquisa_sim = None
        self.radio_pesquisa_nao = None

        # Configurações e widgets principais
        self.selected_button = None
        self.stacked_widget = QStackedWidget(self)
        self.stacked_widget.setStyleSheet("""
            QLabel { font-size: 16px; }
            QCheckBox { font-size: 16px; }
            QLineEdit { font-size: 14px; }
        """)
        self.widgets_map = {}
        
        # Configurações de caminho e inicialização da base de dados
        self.database_path = CONTROLE_DADOS
        self.config = load_config_path_id()
        self.pasta_base = Path(self.config.get('pasta_base', str(Path.home() / 'Desktop')))

        # Consolidador de documentos
        self.consolidador = ConsolidarDocumentos(self.dados, self.icons)
        self.consolidador.status_atualizado.connect(self.atualizar_status)

        # Label de status para mostrar atualizações de consolidação
        self.status_label = QLabel(self)

        # Inicialização da interface gráfica e configuração da UI
        self.setup_ui()

        self.colunas_legiveis = {
            'nup': 'NUP',
            'material_servico': 'Material (M) ou Serviço (S)',
            'vigencia': 'Vigência (Ex: 2 (dois) meses, 6 (seis) meses), etc.',
            'criterio_julgamento': 'Critério de Julgamento (Menor Preço ou Maior Desconto)',
            'com_disputa': 'Com disputa? Sim (S) ou Não (N)',
            'pesquisa_preco': 'Pesquisa Concomitante? Sim (S) ou Não (N)',
            'sigla_om': 'OM',
            'setor_responsavel': 'Setor Responsável',
            'cod_par': 'Código PAR',
            'prioridade_par': 'Prioridade PAR (Necessário, Urgente ou Desejável)',
            'cep': 'CEP',
            'endereco': 'Endereço',
            'email': 'Email',
            'telefone': 'Telefone',
            'dias_para_recebimento': 'Dias para Recebimento',
            'horario_para_recebimento': 'Horário para Recebimento',
            'valor_total': 'Valor Total',
            'acao_interna': 'Ação Interna',
            'fonte_recursos': 'Fonte de Recursos',
            'natureza_despesa': 'Natureza da Despesa',
            'unidade_orcamentaria': 'Unidade Orçamentária',
            'ptres': 'PTRES',
            'atividade_custeio': 'Atividade de Custeio',
            'justificativa': 'Justificativa',
            'comunicacao_padronizada': 'Comunicação Padronizada (CP), Ex: 60-25',
        }
        
        # Dicionário inverso para mapeamento reverso das colunas legíveis
        self.colunas_legiveis_inverso = {v: k for k, v in self.colunas_legiveis.items()}

        # Mapas de normalização de valores para campos específicos
        self.normalizacao_valores = {
            'material_servico': {
                'M': 'Material', 'm': 'Material', 'Material': 'Material', 'material': 'Material',
                'S': 'Serviço', 's': 'Serviço', 'Serviço': 'Serviço', 'serviço': 'Serviço',
                'Servico': 'Serviço', 'servico': 'Serviço'
            },
            'com_disputa': {
                'S': 'Sim', 's': 'Sim', 'Sim': 'Sim', 'sim': 'Sim',
                'N': 'Não', 'n': 'Não', 'Não': 'Não', 'não': 'Não',
                'Nao': 'Não', 'nao': 'Não'
            },
            'pesquisa_preco': {
                'S': 'Sim', 's': 'Sim', 'Sim': 'Sim', 'sim': 'Sim',
                'N': 'Não', 'n': 'Não', 'Não': 'Não', 'não': 'Não',
                'Nao': 'Não', 'nao': 'Não'
            },
            'atividade_custeio': {
                'S': 'Sim', 's': 'Sim', 'Sim': 'Sim', 'sim': 'Sim',
                'N': 'Não', 'n': 'Não', 'Não': 'Não', 'não': 'Não',
                'Nao': 'Não', 'nao': 'Não'
            }
        }            

    def create_agentes_responsaveis_layout(self, max_width=300):
        """Cria o layout para os agentes responsáveis e o organiza em um QGroupBox estilizado."""
        group_box = QGroupBox("Agentes Responsáveis")
        group_box.setMaximumWidth(max_width)
        group_box.setStyleSheet("""
            QGroupBox {
                border: 1px solid #3C3C5A;
                border-radius: 10px;
                font-size: 16px;
                font-weight: bold;
                color: white;
                margin-top: 13px;
            }
            QGroupBox:title {
                subcontrol-origin: margin;
            }
        """)

        # Frame para organizar os agentes responsáveis
        frame_agentes = QFrame()
        agente_responsavel_layout = QVBoxLayout(frame_agentes)

        # Criação dos ComboBox com ajuste de altura
        self.ordenador_combo = create_combo_box('', [], 270, 58)
        self.agente_fiscal_combo = create_combo_box('', [], 270, 58)
        self.gerente_credito_combo = create_combo_box('', [], 270, 58)
        self.responsavel_demanda_combo = create_combo_box('', [], 270, 58)
        self.operador_dispensa_combo = create_combo_box('', [], 270, 58)

        # Adicionando labels e ComboBox diretamente ao layout
        labels_combos = [
            ("Ordenador de Despesa:", self.ordenador_combo),
            ("Agente Fiscal:", self.agente_fiscal_combo),
            ("Gerente de Crédito:", self.gerente_credito_combo),
            ("Responsável pela Demanda:", self.responsavel_demanda_combo),
            ("Operador da Contratação:", self.operador_dispensa_combo)
        ]

        for label_text, combo_box in labels_combos:
            h_layout = QVBoxLayout()
            h_layout.setSpacing(0)
            h_layout.setContentsMargins(0, 0, 0, 0)
            label = QLabel(label_text)
            label.setStyleSheet("color: #8AB4F7; font-size: 16px")
            h_layout.addWidget(label)
            h_layout.addWidget(combo_box)
            agente_responsavel_layout.addLayout(h_layout)

        # Carrega os agentes responsáveis para popular os ComboBoxes
        carregar_agentes_responsaveis(self.database_path, {
            "Ordenador de Despesa%": self.ordenador_combo,
            "Agente Fiscal%": self.agente_fiscal_combo,
            "Gerente de Crédito%": self.gerente_credito_combo,
            "Operador%": self.operador_dispensa_combo,
            "NOT LIKE": self.responsavel_demanda_combo
        })

        # Define o valor inicial de cada ComboBox com os dados passados em `data`
        self.ordenador_combo.setCurrentText(self.dados.get('ordenador_despesas', ''))
        self.agente_fiscal_combo.setCurrentText(self.dados.get('agente_fiscal', ''))
        self.gerente_credito_combo.setCurrentText(self.dados.get('gerente_de_credito', ''))
        self.responsavel_demanda_combo.setCurrentText(self.dados.get('responsavel_pela_demanda', ''))
        self.operador_dispensa_combo.setCurrentText(self.dados.get('operador', ''))

        # Adiciona o frame de agentes ao layout do group_box
        group_layout = QVBoxLayout(group_box)
        group_layout.addWidget(frame_agentes)

        return group_box


    def save_data(self):
        # Coleta os dados dos widgets de contratação
        data_to_save = {
            'id_processo': self.dados.get('id_processo'),
            'tipo': self.dados.get('tipo'),
            'numero': self.dados.get('numero'),
            'ano': self.dados.get('ano'),
            'situacao': self.situacao_combo.currentText(),
            'sigla_om': self.om_combo.currentText(),
            'setor_responsavel': self.setor_responsavel_combo.currentText(),
            'data_sessao': self.data_edit.date().toString("yyyy-MM-dd"),
            'cnpj_matriz': self.cnpj_edit.text(),
            'sequencial_pncp': self.sequencial_edit.text(),
            'objeto': self.objeto_edit.text(),
            'nup': self.nup_edit.text(),
            # Corrige para usar o valor dos RadioButtons
            'material_servico': "Serviço" if self.radio_servico.isChecked() else "Material",
            'vigencia': self.vigencia_combo.currentText(),
            'criterio_julgamento': self.criterio_combo.currentText(),
            'com_disputa': "Sim" if self.radio_disputa_sim.isChecked() else "Não",
            'pesquisa_preco': "Sim" if self.radio_pesquisa_sim.isChecked() else "Não"
        }
        
        # Coleta os dados dos widgets de classificação orçamentária
        data_to_save.update({
            'valor_total': self.valor_total_edit.text(),
            'acao_interna': self.acao_interna_edit.text(),
            'fonte_recursos': self.fonte_recursos_edit.text(),
            'natureza_despesa': self.natureza_despesa_edit.text(),
            'unidade_orcamentaria': self.unidade_orcamentaria_edit.text(),
            'ptres': self.ptres_edit.text(),
            'atividade_custeio': 'Sim' if self.radio_custeio_sim.isChecked() else 'Não'
        })

        data_to_save.update({
            'cp': self.widget_setor_responsavel['cp_edit'].text(),
            'cod_par': self.widget_setor_responsavel['par_edit'].text(),
            'prioridade_par': self.widget_setor_responsavel['prioridade_combo'].currentText(),
            'endereco': self.widget_setor_responsavel['endereco_edit'].text(),
            'cep': self.widget_setor_responsavel['cep_edit'].text(),
            'email': self.widget_setor_responsavel['email_edit'].text(),
            'telefone': self.widget_setor_responsavel['telefone_edit'].text(),
            'dias_recebimento': self.widget_setor_responsavel['dias_edit'].text(),
            'horario_recebimento': self.widget_setor_responsavel['horario_edit'].text(),
            'justificativa': self.widget_setor_responsavel['justificativa_edit'].toPlainText()
        })

        data_to_save.update({
            'ordenador_despesas': self.ordenador_combo.currentText(),
            'agente_fiscal': self.agente_fiscal_combo.currentText(),
            'gerente_de_credito': self.gerente_credito_combo.currentText(),
            'responsavel_pela_demanda': self.responsavel_demanda_combo.currentText(),
            'operador': self.operador_dispensa_combo.currentText()
        })
        
        # Debug para verificar o conteúdo de data_to_save
        print("Dados para salvar:", data_to_save)

        # Emissão do sinal para salvar os dados
        self.save_data_signal.emit(data_to_save)

    def atualizar_status(self, status_texto, icone_path):
        """Atualiza o texto e o ícone do status_label"""
        # Print para verificar a recepção do sinal
        print(f"Received signal in atualizar_status with status_texto: '{status_texto}', icone_path: '{icone_path}'")
        
        self.status_label.setText(status_texto)
        icon_folder = QIcon(icone_path)
        icon_pixmap = icon_folder.pixmap(30, 30)
        self.icon_label.setPixmap(icon_pixmap)  # Atualiza o pixmap do ícone

    def setup_ui(self):
        # Configura o widget principal e define o fundo preto e borda
        main_widget = QWidget(self)
        self.setCentralWidget(main_widget)

        # Configuração do layout principal com margens e espaçamento zero
        self.central_layout = QHBoxLayout(main_widget)

        # Layout esquerdo
        Left_layout = QVBoxLayout()
        layout_titulo = self.setup_layout_titulo()
        Left_layout.addLayout(layout_titulo)
        nav_frame = self.create_navigation_layout()
        Left_layout.addWidget(nav_frame)
        Left_layout.addWidget(self.stacked_widget)
        self.central_layout.addLayout(Left_layout)

        # Configura o layout da consulta API
        self.group_box_consulta_api, self.cnpj_edit, self.sequencial_edit = setup_consulta_api(parent=self, icons=self.icons, data=self.dados)
        self.group_box_agentes = self.create_agentes_responsaveis_layout()
        self.group_box_sessao = self.create_sessao_publica_group()

        # Cria um widget para o Right_layout e define o fundo preto
        right_widget = QWidget()
        right_widget.setStyleSheet("""
            background-color: #12131D;
            border: 2px solid #12131D;
            border-radius: 15px;
        """)
        Right_layout = QVBoxLayout(right_widget)
        Right_layout.addWidget(self.group_box_consulta_api)
        Right_layout.addWidget(self.group_box_agentes)
        Right_layout.addWidget(self.group_box_sessao)

        self.central_layout.addWidget(right_widget)

        # Configuração dos widgets no QStackedWidget
        self.setup_stacked_widgets()

    def verificar_pastas(self, pasta_base):
        # Acesse o id_processo a partir de self.dados
        id_processo = self.dados.get('id_processo', 'desconhecido').replace("/", "-")  # Use uma chave de dicionário
        objeto = self.dados.get('objeto', 'objeto_desconhecido').replace("/", "-")  # Acessando corretamente o objeto

        base_path = pasta_base / f'{id_processo} - {objeto}'

        pastas_necessarias = [
            base_path / '1. Autorizacao',
            base_path / '2. CP e anexos',
            base_path / '3. Aviso',
            base_path / '2. CP e anexos' / 'DFD',
            base_path / '2. CP e anexos' / 'DFD' / 'Anexo A - Relatorio Safin',
            base_path / '2. CP e anexos' / 'DFD' / 'Anexo B - Especificações e Quantidade',
            base_path / '2. CP e anexos' / 'TR',
            base_path / '2. CP e anexos' / 'TR' / 'Pesquisa de Preços',
            base_path / '2. CP e anexos' / 'Declaracao de Adequação Orçamentária',
            base_path / '2. CP e anexos' / 'Declaracao de Adequação Orçamentária' / 'Relatório do PDM-Catser',
            base_path / '2. CP e anexos' / 'Justificativas Relevantes',
        ]

        # Verifica se todas as pastas necessárias existem
        pastas_existentes = all(pasta.exists() for pasta in pastas_necessarias)
        return pastas_existentes

    def create_navigation_layout(self):
        # Criação do frame que conterá o nav_layout e aplicará a borda inferior
        nav_frame = QFrame()
        nav_frame.setStyleSheet("QFrame {border-bottom: 5px solid #13141F;}")

        # Layout horizontal para os botões de navegação
        nav_layout = QHBoxLayout(nav_frame)
        nav_layout.setSpacing(0)
        nav_layout.setContentsMargins(0, 0, 0, 0)

        buttons = [
            ("Informações", "Informações"),
            ("Setor Responsável", "Setor Responsável"),
            ("Documentos", "Documentos"),
            ("Anexos", "Anexos"),
            ("Resultados", "Resultados"),
        ]

        for index, (text, name) in enumerate(buttons):
            button = QPushButton(text, self)
            button.setObjectName(name)
            button.setProperty("class", "nav-button")
            button.clicked.connect(lambda _, n=name, b=button: self.on_navigation_button_clicked(n, b))
            nav_layout.addWidget(button)

            # Define o botão "Informações" como selecionado
            if name == "Informações":
                button.setProperty("class", "nav-button selected")
                button.setStyleSheet("")  # Aplica o estilo imediatamente
                self.selected_button = button  # Mantém o botão "Informações" como o selecionado inicial

        nav_layout.addSpacerItem(QSpacerItem(20, 20, QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Minimum))

        # Verifica se as pastas existem e define o ícone e status
        pastas_existentes = self.verificar_pastas(self.pasta_base)
        status_text = "Pastas encontradas" if pastas_existentes else "Pastas não encontradas"
        icon_key = "folder_v" if pastas_existentes else "folder_x"
        icon = self.icons.get(icon_key)
        self.status_label = QLabel(status_text)
        self.icon_label = QLabel()
        if icon and isinstance(icon, QIcon):
            icon_pixmap = icon.pixmap(30, 30)
            self.icon_label.setPixmap(icon_pixmap)

        # Layout de status com ícone e texto
        status_layout = QHBoxLayout()
        status_layout.addWidget(self.icon_label)
        status_layout.addWidget(self.status_label)
        status_layout.addStretch()
        nav_layout.addLayout(status_layout)

        # Define o estilo para os botões dentro do nav_layout
        self.setStyleSheet("""
            QPushButton[class="nav-button"] {
                background-color: #181928;
                color: #8BE9D9;
                border: none;
                font-size: 14px;
                border-top: 10px solid #181928;
                border-bottom: 5px solid #13141F;                           
            }
            QPushButton[class="nav-button"]:hover {
                background-color: #181928;
                color: white;
                font-size: 14px;
                border-top: 3px solid #181928;
                border-bottom: 5px solid #13141F;
            }
            QPushButton[class="nav-button selected"] {
                background-color: #13141F;
                color: white;
                font-weight: bold;
                font-size: 14px;
                border-top: 2px solid #FF79C6;
                border-bottom: 5px solid #13141F;
            }
        """)

        return nav_frame

    def on_navigation_button_clicked(self, name, button):
        if self.selected_button:
            self.selected_button.setProperty("class", "nav-button")
            self.selected_button.setStyleSheet("")

        button.setProperty("class", "nav-button selected")
        button.setStyleSheet("")

        self.selected_button = button
        self.show_widget(name)

    def show_widget(self, name):
        widget = self.widgets_map.get(name)
        if widget:
            self.stacked_widget.setCurrentWidget(widget)

    def setup_stacked_widgets(self):
        """Configura os widgets para cada seção e os adiciona ao QStackedWidget"""
        # Dados de exemplo para preencher os widgets
        data = self.dados  # Utilize os dados reais

        # Cria widgets para cada seção
        self.widgets_map = {
            "Informações": self.stacked_widget_info(data),
            "Setor Responsável": self.stacked_widget_responsaveis(data),
            "Documentos": self.stacked_widget_documentos(data),
            "Anexos": self.stacked_widget_anexos(data),
            "Resultados": self.stacked_widget_pncp(data),
        }

        # Adiciona cada widget ao QStackedWidget
        for name, widget in self.widgets_map.items():
            self.stacked_widget.addWidget(widget)

    def stacked_widget_info(self, data):
        frame = QFrame()
        layout = QVBoxLayout()
        hbox_top_layout = QHBoxLayout()

        info_contratacao_layout = QVBoxLayout()
        self.contratacao_layout = self.create_contratacao_group()
        info_comprasnet_layout = create_info_comprasnet(data)
        info_contratacao_layout.addWidget(self.contratacao_layout)
        info_contratacao_layout.addWidget(info_comprasnet_layout)

        classificacao_orcamentaria_formulario_layout = QVBoxLayout()
        self.classificacao_orcamentaria_group_box = self.create_classificacao_orcamentaria_group()
        
        self.group_box_formulario = self.setup_formularios()

        classificacao_orcamentaria_formulario_layout.addWidget(self.classificacao_orcamentaria_group_box)
        classificacao_orcamentaria_formulario_layout.addWidget(self.group_box_formulario)

        hbox_top_layout.addLayout(info_contratacao_layout)
        hbox_top_layout.addLayout(classificacao_orcamentaria_formulario_layout)

        layout.addLayout(hbox_top_layout)
        frame.setLayout(layout)

        return frame

    def create_contratacao_group(self):
        contratacao_group_box = QGroupBox("Informações da Contratação")
        contratacao_group_box.setStyleSheet("""
            QGroupBox {
                border: 1px solid #3C3C5A;
                border-radius: 10px;
                font-size: 16px;
                font-weight: bold;
                color: white;
                margin-top: 13px;
            }
            QGroupBox:title {
                subcontrol-origin: margin;
                padding: 0 3px;
            }
        """)

        contratacao_layout = QVBoxLayout()

        # Campo de Objeto
        objeto_layout = QHBoxLayout()
        objeto_label = QLabel("Objeto:")
        self.objeto_edit = QLineEdit(self.dados.get('objeto', ''))
        objeto_layout.addWidget(objeto_label)
        objeto_layout.addWidget(self.objeto_edit)
        contratacao_layout.addLayout(objeto_layout)

        # NUP, Material e Serviço com seleção exclusiva
        nup_layout = QHBoxLayout()
        nup_label = QLabel("NUP:")
        self.nup_edit = QLineEdit(self.dados.get('nup', ''))
        nup_layout.addWidget(nup_label)
        nup_layout.addWidget(self.nup_edit)

        contratacao_layout.addLayout(nup_layout)

        # Layout para Vigência e Critério de Julgamento
        vigencia_layout = QHBoxLayout()

        # Vigência ComboBox
        vigencia_label = QLabel("Vigência:")
        self.vigencia_combo = QComboBox()
        self.vigencia_combo.setEditable(True)
        self.vigencia_combo.setStyleSheet("font-size: 14px;")  # Define o tamanho da fonte via stylesheet
        for i in range(1, 13):
            self.vigencia_combo.addItem(f"{i} ({number_to_text(i)}) meses")
        vigencia = self.dados.get('vigencia', '2 (dois) meses')
        self.vigencia_combo.setCurrentText(vigencia)

        # Expansão horizontal apenas, mantendo altura padrão
        self.vigencia_combo.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)

        # Adiciona Vigência ao layout
        vigencia_layout.addWidget(vigencia_label)
        vigencia_layout.addWidget(self.vigencia_combo)
        contratacao_layout.addLayout(vigencia_layout)
        
        criterio_layout = QHBoxLayout()
        # Critério de Julgamento ComboBox
        criterio_label = QLabel("Critério Julgamento:")
        self.criterio_combo = QComboBox()
        self.criterio_combo.setStyleSheet("font-size: 14px;")  # Define o tamanho da fonte via stylesheet
        self.criterio_combo.addItems(["Menor Preço", "Maior Desconto"])
        self.criterio_combo.setCurrentText(self.dados.get('criterio_julgamento', 'Menor Preço'))

        # Expansão horizontal apenas, mantendo altura padrão
        self.criterio_combo.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)

        # Adiciona Critério de Julgamento ao layout
        criterio_layout.addWidget(criterio_label)
        criterio_layout.addWidget(self.criterio_combo)

        # Adiciona o layout ao layout principal de contratação
        contratacao_layout.addLayout(criterio_layout)

        # Material e Serviço com seleção exclusiva usando RadioButtons
        material_servico_layout = QHBoxLayout()
        material_servico_label = QLabel("Material/Serviço:")
        self.radio_material = QRadioButton("Material")
        self.radio_servico = QRadioButton("Serviço")

        # Grupo de botões exclusivo para Material e Serviço
        self.material_servico_group = QButtonGroup()
        self.material_servico_group.addButton(self.radio_material)
        self.material_servico_group.addButton(self.radio_servico)

        # Define o estado inicial
        material_servico = self.dados.get('material_servico', 'Material')
        self.radio_servico.setChecked(material_servico == "Serviço")
        self.radio_material.setChecked(material_servico == "Material")

        material_servico_layout.addWidget(material_servico_label)
        material_servico_layout.addWidget(self.radio_material)
        material_servico_layout.addWidget(self.radio_servico)
        contratacao_layout.addLayout(material_servico_layout)

        # Configuração dos RadioButtons para "Com Disputa"
        disputa_layout = QHBoxLayout()
        disputa_label = QLabel("Com disputa?")
        self.radio_disputa_sim = QRadioButton("Sim")
        self.radio_disputa_nao = QRadioButton("Não")
        self.disputa_group = QButtonGroup()
        self.disputa_group.addButton(self.radio_disputa_sim)
        self.disputa_group.addButton(self.radio_disputa_nao)

        com_disputa_value = self.dados.get('com_disputa', 'Sim')
        self.radio_disputa_sim.setChecked(com_disputa_value == 'Sim')
        self.radio_disputa_nao.setChecked(com_disputa_value == 'Não')

        disputa_layout.addWidget(disputa_label)
        disputa_layout.addWidget(self.radio_disputa_sim)
        disputa_layout.addWidget(self.radio_disputa_nao)
        contratacao_layout.addLayout(disputa_layout)

        # Configuração dos RadioButtons para "Pesquisa Concomitante"
        pesquisa_layout = QHBoxLayout()
        pesquisa_label = QLabel("Pesquisa Concomitante?")
        self.radio_pesquisa_sim = QRadioButton("Sim")
        self.radio_pesquisa_nao = QRadioButton("Não")
        self.pesquisa_group = QButtonGroup()
        self.pesquisa_group.addButton(self.radio_pesquisa_sim)
        self.pesquisa_group.addButton(self.radio_pesquisa_nao)

        pesquisa_preco_value = self.dados.get('pesquisa_preco', 'Não')
        self.radio_pesquisa_sim.setChecked(pesquisa_preco_value == 'Sim')
        self.radio_pesquisa_nao.setChecked(pesquisa_preco_value == 'Não')

        pesquisa_layout.addWidget(pesquisa_label)
        pesquisa_layout.addWidget(self.radio_pesquisa_sim)
        pesquisa_layout.addWidget(self.radio_pesquisa_nao)
        contratacao_layout.addLayout(pesquisa_layout)

        # Configura layout do GroupBox
        contratacao_group_box.setLayout(contratacao_layout)
        return contratacao_group_box

    def create_classificacao_orcamentaria_group(self):
        """Cria o QGroupBox para a seção de Classificação Orçamentária."""
        classificacao_orcamentaria_group_box = QGroupBox("Classificação Orçamentária")
        classificacao_orcamentaria_group_box.setStyleSheet("""
            QGroupBox {
                border: 1px solid #3C3C5A;
                border-radius: 10px;
                font-size: 16px;
                font-weight: bold;
                color: white;
                margin-top: 13px;
            }
            QGroupBox:title {
                subcontrol-origin: margin;
                padding: 0 3px;
            }
        """)
        classificacao_orcamentaria_group_box.setFixedWidth(400)
        
        layout = QVBoxLayout()

        # Criando cada campo de entrada e armazenando como atributo da classe
        valor_total_layout = QHBoxLayout()
        valor_total_label = QLabel("Valor Estimado:")
        self.valor_total_edit = QLineEdit(str(self.dados.get('valor_total', '')))
        valor_total_layout.addWidget(valor_total_label)
        valor_total_layout.addWidget(self.valor_total_edit)
        layout.addLayout(valor_total_layout)

        acao_interna_layout = QHBoxLayout()
        acao_interna_label = QLabel("Ação Interna:")
        self.acao_interna_edit = QLineEdit(str(self.dados.get('acao_interna', '')))
        acao_interna_layout.addWidget(acao_interna_label)
        acao_interna_layout.addWidget(self.acao_interna_edit)
        layout.addLayout(acao_interna_layout)

        fonte_recursos_layout = QHBoxLayout()
        fonte_recursos_label = QLabel("Fonte de Recurso (FR):")
        self.fonte_recursos_edit = QLineEdit(str(self.dados.get('fonte_recursos', '')))
        fonte_recursos_layout.addWidget(fonte_recursos_label)
        fonte_recursos_layout.addWidget(self.fonte_recursos_edit)
        layout.addLayout(fonte_recursos_layout)

        natureza_despesa_layout = QHBoxLayout()
        natureza_despesa_label = QLabel("Natureza de Despesa (ND):")
        self.natureza_despesa_edit = QLineEdit(str(self.dados.get('natureza_despesa', '')))
        natureza_despesa_layout.addWidget(natureza_despesa_label)
        natureza_despesa_layout.addWidget(self.natureza_despesa_edit)
        layout.addLayout(natureza_despesa_layout)

        unidade_orcamentaria_layout = QHBoxLayout()
        unidade_orcamentaria_label = QLabel("Unidade Orçamentária (UO):")
        self.unidade_orcamentaria_edit = QLineEdit(str(self.dados.get('unidade_orcamentaria', '')))
        unidade_orcamentaria_layout.addWidget(unidade_orcamentaria_label)
        unidade_orcamentaria_layout.addWidget(self.unidade_orcamentaria_edit)
        layout.addLayout(unidade_orcamentaria_layout)

        ptres_layout = QHBoxLayout()
        ptres_label = QLabel("PTRES:")
        self.ptres_edit = QLineEdit(str(self.dados.get('ptres', '')))
        ptres_layout.addWidget(ptres_label)
        ptres_layout.addWidget(self.ptres_edit)
        layout.addLayout(ptres_layout)

        # Adicionando o rádio button de Atividade de Custeio
        custeio_layout = QHBoxLayout()
        custeio_label = QLabel("Atividade de Custeio?")
        self.radio_custeio_sim = QRadioButton("Sim")
        self.radio_custeio_nao = QRadioButton("Não")
        custeio_group = QButtonGroup()  # Grupo exclusivo para o conjunto de botões
        custeio_group.addButton(self.radio_custeio_sim)
        custeio_group.addButton(self.radio_custeio_nao)

        # Define o estado inicial com base nos dados
        atividade_custeio_value = self.dados.get('atividade_custeio', 'Não')
        self.radio_custeio_sim.setChecked(atividade_custeio_value == 'Sim')
        self.radio_custeio_nao.setChecked(atividade_custeio_value == 'Não')

        custeio_layout.addWidget(custeio_label)
        custeio_layout.addWidget(self.radio_custeio_sim)
        custeio_layout.addWidget(self.radio_custeio_nao)
        
        # Adiciona o layout do rádio button ao layout principal
        layout.addLayout(custeio_layout)
        classificacao_orcamentaria_group_box.setLayout(layout)
        
        return classificacao_orcamentaria_group_box

    def stacked_widget_responsaveis(self, data):
        frame = QFrame()
        layout = QVBoxLayout()
        self.setor_responsavel_layout, self.widget_setor_responsavel = create_dados_responsavel_contratacao_group(data)
        layout.addLayout(self.setor_responsavel_layout)
        frame.setLayout(layout)
        return frame

    def stacked_widget_documentos(self, data):
        frame = QFrame()
        layout = QVBoxLayout()

        # Configuração de parâmetros
        nome_pasta = f"{self.consolidador.id_processo.replace('/', '-')}_{self.consolidador.objeto.replace('/', '-')}"
        criar_e_abrir_pasta = self.consolidador.criar_e_abrir_pasta
        alterar_diretorio_base = self.consolidador.alterar_diretorio_base
        editar_modelo = self.consolidador.editar_modelo

        # Chama `create_gerar_documentos_group` sem passar o parâmetro `consolidador`
        self.gerar_documentos = self.consolidador.create_gerar_documentos_group(
            handle_gerar_autorizacao=self.consolidador.gerar_autorizacao,
            handle_gerar_autorizacao_sidgem=None,  # Removido ou definido como None
            handle_gerar_comunicacao_padronizada=self.consolidador.gerar_comunicacao_padronizada,
            handle_gerar_comunicacao_padronizada_sidgem=None,  # Removido ou definido como None
            handle_gerar_aviso_dispensa=self.consolidador.gerar_aviso_dispensa,
            handle_gerar_aviso_dispensa_sidgem=None  # Removido ou definido como None
        )
        self.sigdem_group = create_GrupoSIGDEM(self.dados, self.icons)

        self.utilidade_group = create_utilidades_group(
            pasta_base=self.pasta_base,
            nome_pasta=nome_pasta,
            icons=self.icons,
            criar_e_abrir_pasta=criar_e_abrir_pasta,
            alterar_diretorio_base=alterar_diretorio_base,
            editar_modelo=editar_modelo
        )

        # Adiciona o layout gerado ao layout principal
        layout.addLayout(self.gerar_documentos)
        layout.addWidget(self.sigdem_group)
        layout.addLayout(self.utilidade_group)
        frame.setLayout(layout)
        return frame
    
    def stacked_widget_anexos(self, data):
        frame = QFrame()
        layout = QVBoxLayout()
        label = QLabel("Conteúdo de Anexos")
        layout.addWidget(label)
        frame.setLayout(layout)
        return frame

    def stacked_widget_pncp(self, data):
        frame = QFrame()
        layout = QVBoxLayout()
        label = QLabel("Conteúdo do PNCP")
        layout.addWidget(label)
        frame.setLayout(layout)
        return frame

    def setup_layout_titulo(self):
        """Configura o layout do título com o ID do processo e a seção de consulta API."""
        layout_titulo = QHBoxLayout()

        spacer_left = QSpacerItem(20, 20, QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Minimum)
        layout_titulo.addSpacerItem(spacer_left)

        # Cria um layout vertical para o título e um layout horizontal para ícones e texto
        vlayout_titulo = QVBoxLayout()
        hlayout_titulo = QHBoxLayout()  # Layout horizontal para ícone esquerdo, título e ícone direito


        # Ícone à esquerda
        brasil_icon = QIcon(self.icons.get("brasil_2", None))
        image_label_esquerda = QLabel()
        image_label_esquerda.setAlignment(Qt.AlignmentFlag.AlignCenter)
        image_label_esquerda.setPixmap(brasil_icon.pixmap(30, 30))
        hlayout_titulo.addWidget(image_label_esquerda)

        # Texto do título centralizado
        tipo = self.dados.get("tipo", "N/A")
        numero = self.dados.get("numero", "N/A")
        ano = self.dados.get("ano", "N/A")
        title_label = QLabel(f"{tipo} nº {numero}/{ano}", self)

        # Define o tamanho da fonte para 18 e em negrito
        font = QFont()
        font.setPointSize(18)
        font.setBold(True)
        title_label.setFont(font)
        title_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        hlayout_titulo.addWidget(title_label)

        # Ícone à direita
        acanto_icon = QIcon(self.icons.get("acanto", None))
        image_label_direita = QLabel()
        image_label_direita.setAlignment(Qt.AlignmentFlag.AlignCenter)
        image_label_direita.setPixmap(acanto_icon.pixmap(40, 40))
        hlayout_titulo.addWidget(image_label_direita)

        # Adiciona o layout horizontal ao layout vertical do título
        vlayout_titulo.addLayout(hlayout_titulo)

        # Cria a linha divisória com espaçamento e adiciona ao layout
        linha_divisoria, spacer_baixo_linha = linha_divisoria_layout()
        vlayout_titulo.addWidget(linha_divisoria)
        vlayout_titulo.addSpacerItem(spacer_baixo_linha)

        # Cria um layout horizontal para o campo "Situação"
        situacao_om_setor_layout = QHBoxLayout()
        spacer_situacao = QSpacerItem(20, 20, QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Minimum)
        situacao_om_setor_layout.addSpacerItem(spacer_situacao)

        situacao_label = QLabel("Situação: ")
        situacao_label.setStyleSheet("font-size: 16px; font-weight: bold")
        situacao_om_setor_layout.addWidget(situacao_label)

        # Cria um combobox para a situação
        self.situacao_combo = QComboBox()
        self.situacao_combo.setStyleSheet("font-size: 14px")
        self.situacao_combo.addItems(["Planejamento", "Aprovado", "Sessão Pública", "Homologado", "Empenhado", "Concluído", "Arquivado"])
        self.situacao_combo.setCurrentText(self.dados.get('situacao', 'Planejamento'))
        situacao_om_setor_layout.addWidget(self.situacao_combo)

        om_layout, self.om_combo = create_selecao_om_layout(
            self.database_path,
            dados=self.dados,
            load_sigla_om_callback=load_sigla_om,
            on_om_changed_callback=on_om_changed
        )

        situacao_om_setor_layout.addLayout(om_layout)
        vlayout_titulo.addLayout(situacao_om_setor_layout)

        divisao_layout = QHBoxLayout()
        divisao_label = QLabel("  Divisão: ")
        divisao_label.setStyleSheet("font-size: 16px; font-weight: bold")
        divisao_layout.addWidget(divisao_label)

        # Criando o QComboBox editável
        self.setor_responsavel_combo = QComboBox()
        self.setor_responsavel_combo .setStyleSheet("font-size: 14px")
        # Adicionando as opções ao ComboBox
        divisoes = [
            "Divisão de Abastecimento",
            "Divisão de Finanças",
            "Divisão de Obtenção",
            "Divisão de Pagamento",
            "Divisão de Administração",
            "Divisão de Subsistência"
        ]
        self.setor_responsavel_combo .addItems(divisoes)

        # Definindo o texto atual com base nos dados fornecidos
        self.setor_responsavel_combo .setCurrentText(self.dados.get('setor_responsavel', 'Selecione a Divisão'))
        divisao_layout.addWidget(self.setor_responsavel_combo )

        situacao_om_setor_layout.addLayout(divisao_layout)
            # Espaçador abaixo da linha divisória
        spacer_baixo_linha = QSpacerItem(5, 5, QSizePolicy.Policy.Minimum, QSizePolicy.Policy.Fixed)
        vlayout_titulo.addSpacerItem(spacer_baixo_linha)

        # Adiciona o layout vertical com título e situação ao layout principal
        layout_titulo.addLayout(vlayout_titulo)

        # Espaçador para empurrar o título e o botão "Salvar" para a direita
        spacer_right = QSpacerItem(20, 20, QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Minimum)
        layout_titulo.addSpacerItem(spacer_right)

        # Botão "Salvar" à direita
        add_button_func("Salvar", "confirm", self.save_data, layout_titulo, self.icons, tooltip="Salvar os Dados")

        return layout_titulo

    def create_sessao_publica_group(self):
        # Criação do QGroupBox para a seção Sessão Pública
        self.sessao_groupbox = QGroupBox("Sessão Pública:")
        self.sessao_groupbox.setMaximumWidth(300)
        self.sessao_groupbox.setStyleSheet("""
            QGroupBox {
                border: 1px solid #3C3C5A;
                border-radius: 10px;
                font-size: 16px;
                font-weight: bold;
                color: white;
                margin-top: 13px;
            }
            QGroupBox:title {
                subcontrol-origin: margin;
                padding: 0 3px;
            }
        """)

        group_layout = QHBoxLayout(self.sessao_groupbox)
        
        # Ícone de calendário ao lado do label "Defina a data:"
        calendar_icon = QIcon(self.icons.get("calendar", None))
        icon_label = QLabel()
        icon_label.setPixmap(calendar_icon.pixmap(40, 40))  # Tamanho do ícone ajustado
        
        # Label "Defina a data:"
        date_label = QLabel("Defina a data:")
        date_label.setStyleSheet("font-size: 16px")
        # Layout horizontal para o ícone e o label "Defina a data:"
        top_layout = QHBoxLayout()
        top_layout.addWidget(icon_label)
        top_layout.addWidget(date_label)
        top_layout.addSpacerItem(QSpacerItem(20, 20, QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Minimum))

        group_layout.addLayout(top_layout)
        
        # Configuração do DateEdit com a data inicial
        self.data_edit = QDateEdit()
        self.data_edit.setStyleSheet("font-size: 14px")
        self.data_edit.setCalendarPopup(True)
        data_sessao_str = self.dados.get('data_sessao', '')
        if data_sessao_str:
            self.data_edit.setDate(QDate.fromString(data_sessao_str, "yyyy-MM-dd"))
        else:
            self.data_edit.setDate(QDate.currentDate())

        group_layout.addWidget(self.data_edit)
        
        return self.sessao_groupbox

    def setup_layout_conteudo(self):
        """Configura o layout de conteúdo com StackedWidget e agentes responsáveis."""
        layout_conteudo = QHBoxLayout()
        
        # Layout StackedWidget e ao lado layout agentes responsáveis
        stacked_widget = QStackedWidget(self)
        layout_conteudo.addWidget(stacked_widget)

        # Layout para agentes responsáveis ao lado do StackedWidget
        agentes_responsaveis_layout = QVBoxLayout()
        layout_conteudo.addLayout(agentes_responsaveis_layout)
        
        return layout_conteudo

    def setup_formularios(self):
        """Configura o layout para consulta à API com campos de CNPJ e Sequencial PNCP."""
        group_box = QGroupBox("Formulário")
        layout = QVBoxLayout(group_box)

        # Aplicando o estilo CSS específico ao GroupBox
        group_box.setStyleSheet("""
            QGroupBox {
                border: 1px solid #3C3C5A;
                border-radius: 10px;
                font-size: 16px;
                font-weight: bold;
                color: white;
                margin-top: 13px;
            }
            QGroupBox:title {
                subcontrol-origin: margin;
                padding: 0 3px;
            }
        """)

        # Layout vertical para os botões
        vlayout_botoes = QVBoxLayout()
        
        # Conectar o botão `Novo Formulário` à função `criar_formulario` usando uma função lambda
        add_button_func("   Novo Formulário   ", "excel_down", lambda: self.criar_formulario(), vlayout_botoes, self.icons, tooltip="Gerar o Formulário")
        add_button_func("Importar Formulário", "excel_up", self.on_import_formulario_clicked, vlayout_botoes, self.icons, tooltip="Carregar o Formulário")

        # Adicionar o layout de botões ao layout principal
        layout.addLayout(vlayout_botoes)
        return group_box
    
    def criar_formulario(self):
        # Inicia o worker para criação do formulário em segundo plano
        self.worker = TableCreationWorker(self.dados, self.colunas_legiveis, self.pasta_base)
        self.worker.file_saved.connect(self.abrir_arquivo)  # Conecta o sinal de conclusão
        self.worker.start()  # Inicia a execução em segundo plano

    def abrir_arquivo(self, file_path):
        """Abre o arquivo salvo com o caminho fornecido."""
        if os.name == 'nt':
            os.startfile(file_path)
        elif os.name == 'posix':
            subprocess.call(['open', file_path])
        else:
            subprocess.call(['xdg-open', file_path])

    def on_import_formulario_clicked(self):
        """Ação disparada ao clicar no botão 'Importar Formulário'."""
        self.carregar_formulario()
            
    def carregar_formulario(self):
        """Permite ao usuário selecionar um arquivo XLSX ou ODS e carrega os dados"""
        # Abre a caixa de diálogo para selecionar o arquivo
        file_path, _ = QFileDialog.getOpenFileName(self, "Selecionar Formulário", "", "Arquivos Excel (*.xlsx *.ods)")

        if not file_path:
            return  # Se o usuário cancelar, encerra a função

        try:
            # Verifica a extensão do arquivo para escolher o método de leitura adequado e pula a primeira linha
            if file_path.endswith('.xlsx'):
                data = pd.read_excel(file_path, engine='openpyxl', header=1)
            elif file_path.endswith('.ods'):
                data = pd.read_excel(file_path, engine='odf', header=1)
            else:
                QMessageBox.warning(self, "Formato Inválido", "Por favor, selecione um arquivo .xlsx ou .ods.")
                return

            # Exibe os dados no console para verificar o conteúdo
            print("Dados do formulário carregados:")
            print(data)

            # Armazena o DataFrame para ser usado na função de preenchimento
            self.df_registro_selecionado = data

            # Preenche os campos específicos com os valores carregados
            self.preencher_dados(data)

        except Exception as e:
            QMessageBox.critical(self, "Erro ao Carregar Formulário", f"Erro ao carregar o formulário: {e}")


    def preencher_dados(self, dados_novos):
        """Preenche os campos da interface com os dados carregados"""
        try:
            def obter_valor(indice):
                """Função auxiliar para obter um valor de 'Valor' com base no 'Índice'"""
                resultado = dados_novos.loc[dados_novos['Índice'] == indice, 'Valor']
                return str(resultado.values[0]) if not resultado.empty else ""

            # Atualizar campos específicos com base nos valores da coluna 'Valor' do DataFrame
            self.valor_total_edit.setText(obter_valor('Valor Total'))
            self.acao_interna_edit.setText(obter_valor('Ação Interna'))
            self.fonte_recursos_edit.setText(obter_valor('Fonte de Recursos'))
            self.natureza_despesa_edit.setText(obter_valor('Natureza da Despesa'))
            self.unidade_orcamentaria_edit.setText(obter_valor('Unidade Orçamentária'))
            self.ptres_edit.setText(obter_valor('PTRES'))
            
            # Configurar o estado do botão de rádio com base no valor encontrado
            atividade_custeio_valor = obter_valor('Atividade de Custeio')
            self.radio_custeio_sim.setChecked(atividade_custeio_valor == 'Sim')
            self.radio_custeio_nao.setChecked(atividade_custeio_valor == 'Não')

        except Exception as e:
            print(f"Erro ao preencher os dados: {str(e)}")
            QMessageBox.critical(self, "Erro", f"Falha ao preencher os dados: {str(e)}")